import sqlite3
from datetime import datetime, timedelta
import json
from openpyxl import load_workbook 
import calendar
import time

def find_index(sheet, column_count, orario_biblio):
    for i in range(3, column_count + 1):
        if sheet.cell(row = 1, column = i).value == orario_biblio:
            return i
    return column_count

def find_index_now(sheet, index_apertura, index_chiusura, data_attuale):
    for i in range(index_apertura, index_chiusura):
        if datetime.strptime(sheet.cell(row=1, column=i).value, '%H:%M') > data_attuale:
            return i - 1
    return index_chiusura

def select_aula(nome_biblio, apertura_biblio="09:00", chiusura_biblio="18:00"):
    giorno = calendar.day_name[datetime.now().weekday()]
    # print("Nome: " + str(nome_biblio))
    # print("Giorno: " + str(giorno))
    # print("Apertura: " + str(apertura_biblio))
    # print("Chiusura: " + str(chiusura_biblio))
    
    try:
        wb = load_workbook('excel/'+ nome_biblio +'/settimana_' + nome_biblio + '.xlsx') 
        sheet = wb[giorno] 
    except:
        print("INFO: impossibile aprire il file [ " + 'excel/'+ nome_biblio +'/settimana_' + nome_biblio + '.xlsx' + " ], non esiste o non esiste il foglio del giorno.")
        return "N/A", "N/A", "N/A", "N/A"

    row_count = sheet.max_row 
    column_count = sheet.max_column
    # print("Row: " + str(row_count))
    # print("Column: " + str(column_count))

    data_attuale = datetime.strptime(str(datetime.now())[11:16], '%H:%M')
    # print("Data: " + str(data))

    index_apertura = find_index(sheet, column_count, apertura_biblio)
    index_chiusura = find_index(sheet, column_count, chiusura_biblio)
    # print("Index apertura: " + str(index_apertura) + " index chiusura: " + str(index_chiusura))

    index_adesso = find_index_now(sheet, index_apertura, index_chiusura, data_attuale)
    # print("Index adesso: " + str(index_adesso))
    
    diz = {}
    for i in range(2, row_count + 1): 
        diz[i] = 0
        
        for j in range(index_adesso, index_chiusura): 
            val = sheet.cell(row = i, column = j).value 
            if val is None:
                diz[i] = diz[i] + 1
            else:
                #print("BREAK")
                break #aula occupata dall'orario necessario per l'apertura
    
    #print(diz.items())
    ris = 0 # number of free time slot
    best_row = 0
    best_cap = 0
    for key, value in diz.items():
        capacity = sheet.cell(row = key, column = 2).value
        if (value == ris and capacity > best_cap and value != 0) or value > ris :
            ris = value
            best_row = key
            best_cap = capacity

    #print(str(ris) + " " + str(best_cap)+ " " + str(best_row))

    if ris < 4: # mettere ris < 4 cosi che le aule le si pare per almeno un'ora
        print("INFO: non esistono aule libere per estendere la biblioteca.")
        return "N/A", "N/A", "N/A", "N/A"
    else:
        apertura = sheet.cell(row = 1, column = index_adesso).value
        chiusura = sheet.cell(row = 1, column = index_adesso + ris).value
        nome_aula = sheet.cell(row = best_row, column = 1).value
        #print("apro aula " + nome_aula + " con capienza : " + str(best_cap) + " dalle " + apertura + " alle " + chiusura )
        return nome_aula, best_cap, apertura, chiusura

# BE CAREFUL: NOT SAFE FROM SQL INJECTION!
def select_from_db(what, table, db_name):
    con = sqlite3.connect(db_name)
    cur = con.cursor()
    
    rows = []
    for row in cur.execute(f"SELECT {what} FROM {table}"):
        rows.append(row)
    
    con.close()
    return rows

def collect_biblioteche():
    return select_from_db("*", "biblioteche", "../tessere.db")

# BE CAREFUL: NOT SAFE FROM SQL INJECTION!
def update_db(table, value, condition, db_name):
    con = sqlite3.connect(db_name)
    cur = con.cursor()

    #cur.execute(f"UPDATE biblioteche SET is_extended = 1 WHERE nome = 'ingmo'")
    cur.execute(f"UPDATE {table} SET {value} WHERE {condition}")
    
    con.commit()
    con.close()

def extend_biblio(nome_biblio, orario_apertura):

    aula, capienza, apertura, chiusura = select_aula(nome_biblio, orario_apertura[0:5], orario_apertura[6:])
    diz = { 
        "name" : aula,
        "capacity" : capienza,
        "open_from" : apertura,
        "open_until" : chiusura
    }
    
    if diz["name"] != "N/A":
        print("INFO: trovato aula per estendere biblioteca!")
        update_db("biblioteche", f"extension = ('{json.dumps(diz)}')", f"nome = '{nome_biblio}'", "../tessere.db")
        update_db("biblioteche", "is_extended = 1", f"nome = '{nome_biblio}'", "../tessere.db")
    
def close_biblio(nome_biblio):
    update_db("biblioteche", f"extension = ('closed')", f"nome = '{nome_biblio}'", "../tessere.db")
    update_db("biblioteche", "is_extended = 0", f"nome = '{nome_biblio}'", "../tessere.db")


def main():
    starttime = time.time()
    
    while True:
        if  (time.time() - starttime) > 60: 
            biblioteche = collect_biblioteche() 
            
            for biblioteca in biblioteche:
                nome = biblioteca[0]
                count = biblioteca[1]
                capienza = biblioteca[2]
                is_extended = biblioteca[3]
                opening_hours = json.loads(biblioteca[5]) if biblioteca[5] != "N/A" else "N/A"
                soglia = capienza - count

                if is_extended == True:
                    extension = json.loads(biblioteca[4])
                    if datetime.strptime(extension["open_until"], "%H:%M") < datetime.strptime(str(datetime.now())[11:16], '%H:%M'):
                        print("Sto chiudendo l'estensione di biblio " + nome)
                        close_biblio(nome)
                
                if soglia <= 2 and is_extended == False:
                    if opening_hours[calendar.day_name[datetime.now().weekday()]] != "N/A":
                        print("Provo ad estendere biblio " + nome)
                        extend_biblio(nome, opening_hours[calendar.day_name[datetime.now().weekday()]])
            starttime = time.time()


if __name__ == "__main__":
    main()